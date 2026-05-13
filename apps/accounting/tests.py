"""Smoke tests for the accounting MVP.

Covers:
- `Transaction.signed_amount` arithmetic (+IN / -OUT).
- `recompute_period_balances` flagging discrepancies correctly.
- Daybook ordering by date.
- XLSX export produces a valid workbook with the expected sheets.
- Setup-wizard gating redirects to `/accounting/setup/` until the prerequisites
  (base currency on the tenant + at least one Account) are configured.
"""
from __future__ import annotations

import datetime as dt
from decimal import Decimal
from io import BytesIO

from django.test import RequestFactory
from django_tenants.test.cases import FastTenantTestCase
from openpyxl import load_workbook

from apps.accounting.models import FiscalPeriod, PeriodAccountBalance
from apps.accounting.services.balance import (
    confirm_opening_balance,
    recompute_period_balances,
)
from apps.accounting.services.daybook import build_daybook
from apps.accounting.services.exports import build_monthly_workbook, build_yearly_workbook
from apps.accounting.services.period import get_or_create_period
from apps.accounting.services.period_feed import build_period_feed
from apps.accounting.setup import is_setup_completed
from apps.money.models import Account, Currency, InvoiceSettlementAllocation, Transaction


class AccountingSmokeTests(FastTenantTestCase):
    @classmethod
    def setup_tenant(cls, tenant):
        from apps.users.models import TenantUser

        owner, _ = TenantUser.objects.get_or_create(
            email="accounting-test@example.com",
            defaults={"is_active": True},
        )
        tenant.name = "Accounting Test Tenant"
        tenant.slug = "accounting-test"
        tenant.owner = owner

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._eur = Currency.objects.get(code="EUR")

    def setUp(self):
        super().setUp()
        Transaction.objects.all().delete()
        PeriodAccountBalance.objects.all().delete()
        FiscalPeriod.objects.all().delete()
        Account.objects.all().delete()

        self.account = Account.objects.create(
            name="Main bank",
            kind=Account.KIND_BANK,
            currency=self._eur,
            opening_balance=Decimal("1000.00"),
            opening_date=dt.date(2026, 4, 30),
        )

    def _make_tx(self, **kwargs):
        invoice = kwargs.pop("invoice", None)
        defaults = {
            "date": dt.date(2026, 5, 5),
            "currency": self._eur,
            "account": self.account,
            "amount": Decimal("100.00"),
            "direction": Transaction.DIRECTION_IN,
        }
        defaults.update(kwargs)
        tx = Transaction.objects.create(**defaults)
        if invoice is not None:
            InvoiceSettlementAllocation.objects.create(
                transaction=tx,
                invoice=invoice,
                amount_settlement=tx.amount,
                amount_invoice=tx.amount,
            )
        return tx

    # ---------------- signed amount ----------------

    def test_signed_amount_is_positive_for_income(self):
        tx = self._make_tx(direction=Transaction.DIRECTION_IN, amount=Decimal("250"))
        self.assertEqual(tx.signed_amount, Decimal("250"))

    def test_signed_amount_is_negative_for_expense(self):
        tx = self._make_tx(direction=Transaction.DIRECTION_OUT, amount=Decimal("80"))
        self.assertEqual(tx.signed_amount, Decimal("-80"))

    # ---------------- daybook ordering ----------------

    def test_daybook_orders_by_date(self):
        self._make_tx(date=dt.date(2026, 5, 10), amount=Decimal("10"))
        self._make_tx(date=dt.date(2026, 5, 1), amount=Decimal("20"))
        self._make_tx(date=dt.date(2026, 5, 5), amount=Decimal("30"))
        rows = list(
            build_daybook(date_from=dt.date(2026, 5, 1), date_to=dt.date(2026, 5, 31))
        )
        dates = [r.date for r in rows]
        self.assertEqual(dates, sorted(dates))

    # ---------------- balance reconciliation ----------------

    def test_recompute_period_balances_flags_mismatch(self):
        period = get_or_create_period(2026, 5)
        self._make_tx(
            date=dt.date(2026, 5, 5),
            amount=Decimal("400"),
            direction=Transaction.DIRECTION_IN,
        )
        self._make_tx(
            date=dt.date(2026, 5, 10),
            amount=Decimal("100"),
            direction=Transaction.DIRECTION_OUT,
        )
        bal = PeriodAccountBalance.objects.create(
            period=period,
            account=self.account,
            starting_balance=Decimal("1000"),
            ending_balance=Decimal("1500"),
        )
        recompute_period_balances(period)
        bal.refresh_from_db()
        self.assertEqual(bal.computed_flow, Decimal("300"))
        self.assertEqual(bal.computed_ending, Decimal("1300"))
        self.assertEqual(bal.discrepancy, Decimal("200"))
        self.assertFalse(bal.is_balanced)

    def test_recompute_period_balances_flags_match(self):
        period = get_or_create_period(2026, 5)
        self._make_tx(
            date=dt.date(2026, 5, 5),
            amount=Decimal("400"),
            direction=Transaction.DIRECTION_IN,
        )
        self._make_tx(
            date=dt.date(2026, 5, 10),
            amount=Decimal("100"),
            direction=Transaction.DIRECTION_OUT,
        )
        PeriodAccountBalance.objects.create(
            period=period,
            account=self.account,
            starting_balance=Decimal("1000"),
            ending_balance=Decimal("1300"),
        )
        recompute_period_balances(period)
        bal = PeriodAccountBalance.objects.get(period=period, account=self.account)
        self.assertEqual(bal.discrepancy, Decimal("0"))
        self.assertTrue(bal.is_balanced)

    # ---------------- xlsx export ----------------

    def test_monthly_workbook_has_expected_sheets(self):
        period = get_or_create_period(2026, 5)
        self._make_tx(date=dt.date(2026, 5, 3), amount=Decimal("123.45"))
        buffer: BytesIO = build_monthly_workbook(period)
        wb = load_workbook(buffer)
        self.assertEqual(
            wb.sheetnames,
            [
                "Daybook",
                "Income & Expenses",
                "Balance",
                "Issued Invoices",
                "Received Invoices",
                "Bank Lines",
            ],
        )

    def test_yearly_workbook_same_sheets_and_month_column_on_daybook(self):
        self._make_tx(date=dt.date(2026, 3, 15), amount=Decimal("50"))
        self._make_tx(date=dt.date(2026, 11, 20), amount=Decimal("25"))
        buffer: BytesIO = build_yearly_workbook(2026)
        wb = load_workbook(buffer)
        self.assertEqual(
            wb.sheetnames,
            [
                "Daybook",
                "Income & Expenses",
                "Balance",
                "Issued Invoices",
                "Received Invoices",
                "Bank Lines",
            ],
        )
        daybook = wb["Daybook"]
        headers = [daybook.cell(row=1, column=c).value for c in range(1, 12)]
        self.assertEqual(headers[1], "Month")

    # ---------------- period feed (daybook admin view) ----------------

    def test_build_period_feed_includes_transactions_and_unpaid_invoices(self):
        from apps.invoices.models import Invoice

        period = get_or_create_period(2026, 5)
        tx_paid_invoice = Invoice.objects.create(
            invoice_number="INV-1",
            invoice_date=dt.date(2026, 5, 4),
            due_date=dt.date(2026, 5, 30),
            total_amount=Decimal("500"),
            currency=self._eur,
            original_filename="inv1.pdf",
            file_path="invoices/inv1.pdf",
        )
        Invoice.objects.create(
            invoice_number="INV-2",
            invoice_date=dt.date(2026, 5, 6),
            due_date=dt.date(2026, 5, 30),
            total_amount=Decimal("250"),
            currency=self._eur,
            original_filename="inv2.pdf",
            file_path="invoices/inv2.pdf",
        )
        self._make_tx(
            date=dt.date(2026, 5, 5),
            amount=Decimal("500"),
            direction=Transaction.DIRECTION_IN,
            invoice=tx_paid_invoice,
        )
        self._make_tx(
            date=dt.date(2026, 5, 1),
            amount=Decimal("80"),
            direction=Transaction.DIRECTION_OUT,
        )

        feed = build_period_feed(period)

        self.assertEqual(feed.period, period)
        self.assertEqual(len(feed.entries), 3)
        self.assertEqual(
            [e["date"] for e in feed.entries],
            sorted([e["date"] for e in feed.entries]),
        )
        kinds = [(e["kind"], e["account_label"]) for e in feed.entries]
        self.assertIn(("invoice", "Unpaid"), kinds)
        self.assertIn(("transaction", self.account.name), kinds)
        self.assertEqual(feed.totals["unpaid_invoice_count"], Decimal("1"))

    # ---------------- confirm_opening_balance ----------------

    def test_confirm_opening_balance_stamps_audit_fields_and_recomputes(self):
        from apps.users.models import TenantUser

        user, _ = TenantUser.objects.get_or_create(
            email="confirmer@example.com",
            defaults={"is_active": True},
        )
        period = get_or_create_period(2026, 5)
        self._make_tx(
            date=dt.date(2026, 5, 4),
            amount=Decimal("200"),
            direction=Transaction.DIRECTION_IN,
        )

        bal = confirm_opening_balance(
            period,
            self.account,
            user,
            starting_balance=Decimal("750"),
        )

        bal.refresh_from_db()
        self.assertEqual(bal.starting_balance, Decimal("750"))
        self.assertIsNotNone(bal.opening_confirmed_at)
        self.assertEqual(bal.opening_confirmed_by_id, user.id)
        self.assertEqual(bal.computed_flow, Decimal("200"))
        self.assertEqual(bal.computed_ending, Decimal("950"))

    def test_recompute_does_not_overwrite_confirmed_starting_balance(self):
        from apps.users.models import TenantUser

        user, _ = TenantUser.objects.get_or_create(
            email="confirmer2@example.com",
            defaults={"is_active": True},
        )
        period = get_or_create_period(2026, 5)
        confirm_opening_balance(
            period,
            self.account,
            user,
            starting_balance=Decimal("0"),
        )
        recompute_period_balances(period)
        bal = PeriodAccountBalance.objects.get(period=period, account=self.account)
        self.assertEqual(bal.starting_balance, Decimal("0"))

    # ---------------- setup wizard gating ----------------

    def test_is_setup_completed_returns_false_when_no_account(self):
        Account.objects.all().delete()
        request = RequestFactory().get("/accounting/")
        request.tenant = self.tenant
        request.tenant.settings_json = {
            "accounting_setup_completed_at": "2026-01-01T00:00:00"
        }
        self.assertFalse(is_setup_completed(request))

    def test_is_setup_completed_returns_true_when_flag_and_account_present(self):
        request = RequestFactory().get("/accounting/")
        request.tenant = self.tenant
        request.tenant.settings_json = {
            "accounting_setup_completed_at": "2026-01-01T00:00:00"
        }
        self.assertTrue(is_setup_completed(request))
