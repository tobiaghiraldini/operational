"""Generate the monthly or yearly XLS package the accountant receives.

Sheets, in order:

1. Daybook       chronological cash/bank flows
2. Income & Expenses    totals by category
3. Balance       per-account opening/flow/closing/discrepancy
4. Issued Invoices
5. Received Invoices
6. Bank Statement Lines (matched / unmatched)

Uses `openpyxl` directly so we can apply header styling, frozen panes, and
currency formats without bringing in a heavier dependency.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any

from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.accounting.models import (
    AccountingExport,
    FiscalPeriod,
    PeriodAccountBalance,
)
from apps.accounting.services.balance import recompute_period_balances
from apps.accounting.services.daybook import build_daybook
from apps.accounting.services.income_statement import (
    build_income_statement,
    build_income_statement_for_range,
)
from apps.accounting.services.period import (
    calendar_year_bounds,
    period_date_range,
)
from apps.invoices.models import Invoice
from apps.money.models import Account, BankStatementLine, Transaction
from apps.money.services.invoice_document_label import transaction_invoice_document_label


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
ZEBRA_FILL = PatternFill("solid", fgColor="F9FAFB")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")

CURRENCY_FORMAT = '#,##0.00;[Red]-#,##0.00'


def build_monthly_workbook(period: FiscalPeriod) -> BytesIO:
    """Return an in-memory XLSX `BytesIO` for `period`."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    date_from, date_to = period_date_range(period)
    _add_daybook_sheet_for_range(workbook, date_from, date_to, include_month=False)
    _add_income_statement_sheet_for_range(workbook, date_from, date_to)
    _add_balance_sheet_for_period(workbook, period)
    _add_invoices_sheet_for_range(
        workbook, date_from, date_to, kind="emitted", title="Issued Invoices"
    )
    _add_invoices_sheet_for_range(
        workbook, date_from, date_to, kind="received", title="Received Invoices"
    )
    _add_bank_lines_sheet_for_range(workbook, date_from, date_to)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_yearly_workbook(year: int) -> BytesIO:
    """Return an XLSX covering Jan 1–Dec 31 of `year` (calendar year)."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    date_from, date_to = calendar_year_bounds(year)
    _add_daybook_sheet_for_range(workbook, date_from, date_to, include_month=True)
    _add_income_statement_sheet_for_range(workbook, date_from, date_to)
    _add_balance_sheet_for_year(workbook, year)
    _add_invoices_sheet_for_range(
        workbook, date_from, date_to, kind="emitted", title="Issued Invoices"
    )
    _add_invoices_sheet_for_range(
        workbook, date_from, date_to, kind="received", title="Received Invoices"
    )
    _add_bank_lines_sheet_for_range(workbook, date_from, date_to)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_workbook_response(
    period: FiscalPeriod, *, user=None
) -> HttpResponse:
    """Build the workbook, persist an `AccountingExport` audit row, return HttpResponse."""
    buffer = build_monthly_workbook(period)
    file_name = f"accounting_{period.label}.xlsx"

    AccountingExport.objects.create(
        period=period,
        kind=AccountingExport.KIND_FULL_PACKAGE,
        file_name=file_name,
        generated_by=user,
        parameters_json={"period": period.label},
    )

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response


def build_yearly_workbook_response(year: int, *, user=None) -> HttpResponse:
    """Year-end package for accountants (single file, all months)."""
    buffer = build_yearly_workbook(year)
    file_name = f"accounting_{year}_full_year.xlsx"

    AccountingExport.objects.create(
        period=None,
        kind=AccountingExport.KIND_FULL_YEAR_PACKAGE,
        file_name=file_name,
        generated_by=user,
        parameters_json={"year": year, "scope": "calendar_year"},
    )

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _add_daybook_sheet_for_range(
    workbook: Workbook,
    date_from,
    date_to,
    *,
    include_month: bool,
) -> None:
    sheet = workbook.create_sheet("Daybook")
    if include_month:
        headers = [
            "Date",
            "Month",
            "Doc #",
            "Counterparty",
            "Description",
            "Account",
            "Category",
            "In",
            "Out",
            "Currency",
            "Balance",
        ]
        in_col, out_col, bal_col = 8, 9, 11
    else:
        headers = [
            "Date",
            "Doc #",
            "Counterparty",
            "Description",
            "Account",
            "Category",
            "In",
            "Out",
            "Currency",
            "Balance",
        ]
        in_col, out_col, bal_col = 7, 8, 10
    _write_header(sheet, headers)

    transactions = build_daybook(date_from=date_from, date_to=date_to)

    balance = Decimal("0")
    for row_idx, tx in enumerate(transactions, start=2):
        balance += tx.signed_amount
        in_amount = tx.amount if tx.direction == Transaction.DIRECTION_IN else None
        out_amount = tx.amount if tx.direction == Transaction.DIRECTION_OUT else None

        month_label = f"{tx.date.year:04d}-{tx.date.month:02d}"
        base_row = [
            tx.date,
            transaction_invoice_document_label(tx),
            tx.counterparty
            or (tx.customer.name if tx.customer else "")
            or (tx.vendor.name if tx.vendor else ""),
            tx.description,
            tx.account.name if tx.account else "",
            tx.category.name if tx.category else "",
            in_amount,
            out_amount,
            tx.currency.code if tx.currency else "",
            balance,
        ]
        if include_month:
            row = [base_row[0], month_label] + base_row[1:]
        else:
            row = base_row

        _write_row(sheet, row_idx, row, zebra=row_idx % 2 == 0)
        for col in (in_col, out_col, bal_col):
            sheet.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT

    _autosize(sheet, headers)
    sheet.freeze_panes = "A2"


def _add_income_statement_sheet_for_range(workbook: Workbook, date_from, date_to) -> None:
    sheet = workbook.create_sheet("Income & Expenses")
    headers = ["Category", "Income", "Expense", "Net"]
    _write_header(sheet, headers)

    summary = build_income_statement_for_range(date_from, date_to)
    income_lookup = {row["category_id"]: row for row in summary["income_by_category"]}
    expense_lookup = {row["category_id"]: row for row in summary["expense_by_category"]}
    keys = list({*income_lookup.keys(), *expense_lookup.keys()})

    row_idx = 2
    for cat_id in keys:
        income_row = income_lookup.get(cat_id, {"name": "Uncategorized", "total": Decimal("0")})
        expense_row = expense_lookup.get(cat_id, {"name": income_row.get("name", "Uncategorized"), "total": Decimal("0")})
        name = income_row.get("name") or expense_row.get("name") or "Uncategorized"
        income = income_row.get("total", Decimal("0"))
        expense = expense_row.get("total", Decimal("0"))
        _write_row(
            sheet,
            row_idx,
            [name, income, expense, income - expense],
            zebra=row_idx % 2 == 0,
        )
        for col in (2, 3, 4):
            sheet.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT
        row_idx += 1

    total_row = [
        "TOTAL",
        summary["total_income"],
        summary["total_expense"],
        summary["net"],
    ]
    _write_row(sheet, row_idx, total_row, bold=True)
    for col in (2, 3, 4):
        sheet.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT

    _autosize(sheet, headers)
    sheet.freeze_panes = "A2"


def _add_balance_sheet_for_period(workbook: Workbook, period: FiscalPeriod) -> None:
    sheet = workbook.create_sheet("Balance")
    headers = [
        "Account",
        "Currency",
        "Opening",
        "Total In",
        "Total Out",
        "Computed Closing",
        "Reported Closing",
        "Discrepancy",
        "OK?",
    ]
    _write_header(sheet, headers)

    recompute_period_balances(period)
    rows = (
        PeriodAccountBalance.objects.filter(period=period)
        .select_related("account", "account__currency")
        .order_by("account__name")
    )
    date_from, date_to = period_date_range(period)

    row_idx = 2
    for bal in rows:
        in_total = _sum_amount_for_direction(
            bal.account, Transaction.DIRECTION_IN, date_from, date_to
        )
        out_total = _sum_amount_for_direction(
            bal.account, Transaction.DIRECTION_OUT, date_from, date_to
        )
        ok_label = "Yes" if bal.is_balanced else "No"
        _write_row(
            sheet,
            row_idx,
            [
                bal.account.name,
                bal.account.currency.code if bal.account.currency_id else "",
                bal.starting_balance,
                in_total,
                out_total,
                bal.computed_ending,
                bal.ending_balance,
                bal.discrepancy,
                ok_label,
            ],
            zebra=row_idx % 2 == 0,
        )
        for col in (3, 4, 5, 6, 7, 8):
            sheet.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT
        ok_cell = sheet.cell(row=row_idx, column=9)
        ok_cell.fill = GREEN_FILL if bal.is_balanced else RED_FILL
        row_idx += 1

    _autosize(sheet, headers)
    sheet.freeze_panes = "A2"


def _add_balance_sheet_for_year(workbook: Workbook, year: int) -> None:
    """Per-account roll-up: opening from January period balance (or account default),
    full-year flows, reported closing from December when present."""
    sheet = workbook.create_sheet("Balance")
    headers = [
        "Account",
        "Currency",
        "Opening",
        "Total In",
        "Total Out",
        "Computed Closing",
        "Reported Closing",
        "Discrepancy",
        "OK?",
    ]
    _write_header(sheet, headers)

    date_from, date_to = calendar_year_bounds(year)
    jan_period = FiscalPeriod.objects.filter(year=year, month=1).first()
    dec_period = FiscalPeriod.objects.filter(year=year, month=12).first()

    row_idx = 2
    for account in Account.objects.filter(is_active=True).order_by("name"):
        jan_bal = None
        if jan_period:
            jan_bal = PeriodAccountBalance.objects.filter(
                period=jan_period, account=account
            ).first()
        opening = (
            jan_bal.starting_balance
            if jan_bal is not None
            else (account.opening_balance or Decimal("0"))
        )

        in_total = _sum_amount_for_direction(
            account, Transaction.DIRECTION_IN, date_from, date_to
        )
        out_total = _sum_amount_for_direction(
            account, Transaction.DIRECTION_OUT, date_from, date_to
        )
        computed_flow = in_total - out_total
        computed_ending = opening + computed_flow

        dec_bal = None
        if dec_period:
            dec_bal = PeriodAccountBalance.objects.filter(
                period=dec_period, account=account
            ).first()
        reported = dec_bal.ending_balance if dec_bal is not None else Decimal("0")

        discrepancy = reported - computed_ending
        is_balanced = discrepancy == Decimal("0") and reported != Decimal("0")
        ok_label = "Yes" if is_balanced else "No"

        _write_row(
            sheet,
            row_idx,
            [
                account.name,
                account.currency.code if account.currency_id else "",
                opening,
                in_total,
                out_total,
                computed_ending,
                reported,
                discrepancy,
                ok_label,
            ],
            zebra=row_idx % 2 == 0,
        )
        for col in (3, 4, 5, 6, 7, 8):
            sheet.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT
        ok_cell = sheet.cell(row=row_idx, column=9)
        ok_cell.fill = GREEN_FILL if is_balanced else RED_FILL
        row_idx += 1

    _autosize(sheet, headers)
    sheet.freeze_panes = "A2"


def _sum_amount_for_direction(
    account: Account, direction: str, date_from, date_to
) -> Decimal:
    total = (
        Transaction.objects.filter(
            account=account,
            direction=direction,
            date__gte=date_from,
            date__lte=date_to,
        ).aggregate(total=Sum("amount"))["total"]
        or Decimal("0")
    )
    return total


def _add_invoices_sheet_for_range(
    workbook: Workbook,
    date_from,
    date_to,
    *,
    kind: str,
    title: str,
) -> None:
    sheet = workbook.create_sheet(title)
    headers = [
        "Date",
        "#",
        "Counterparty",
        "Net",
        "VAT",
        "Gross",
        "Currency",
        "Status",
        "Paid On",
    ]
    _write_header(sheet, headers)

    invoices = (
        Invoice.objects.filter(
            invoice_type=kind,
            invoice_date__gte=date_from,
            invoice_date__lte=date_to,
        )
        .select_related("customer", "vendor", "currency")
        .order_by("invoice_date", "id")
    )
    for row_idx, invoice in enumerate(invoices, start=2):
        counterparty = ""
        if kind == "emitted" and invoice.customer:
            counterparty = invoice.customer.name
        elif kind == "received" and invoice.vendor:
            counterparty = invoice.vendor.name
        _write_row(
            sheet,
            row_idx,
            [
                invoice.invoice_date,
                invoice.invoice_number,
                counterparty,
                invoice.taxable_amount,
                invoice.vat_amount,
                invoice.total_amount,
                invoice.currency.code if invoice.currency_id else "",
                invoice.status,
                invoice.payment_date,
            ],
            zebra=row_idx % 2 == 0,
        )
        for col in (4, 5, 6):
            sheet.cell(row=row_idx, column=col).number_format = CURRENCY_FORMAT

    _autosize(sheet, headers)
    sheet.freeze_panes = "A2"


def _add_bank_lines_sheet_for_range(workbook: Workbook, date_from, date_to) -> None:
    sheet = workbook.create_sheet("Bank Lines")
    headers = ["Date", "Description", "Direction", "Amount", "Matched?", "Linked Tx"]
    _write_header(sheet, headers)

    lines = (
        BankStatementLine.objects.filter(date__gte=date_from, date__lte=date_to)
        .select_related("statement", "statement__account", "matched_transaction")
        .order_by("date", "id")
    )
    for row_idx, line in enumerate(lines, start=2):
        _write_row(
            sheet,
            row_idx,
            [
                line.date,
                line.description,
                line.direction,
                line.amount,
                "Yes" if line.is_matched else "No",
                str(line.matched_transaction) if line.matched_transaction else "",
            ],
            zebra=row_idx % 2 == 0,
        )
        sheet.cell(row=row_idx, column=4).number_format = CURRENCY_FORMAT
        match_cell = sheet.cell(row=row_idx, column=5)
        match_cell.fill = GREEN_FILL if line.is_matched else RED_FILL

    _autosize(sheet, headers)
    sheet.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_row(
    sheet: Worksheet,
    row_idx: int,
    values: list[Any],
    *,
    zebra: bool = False,
    bold: bool = False,
) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        if zebra:
            cell.fill = ZEBRA_FILL
        if bold:
            cell.font = Font(bold=True)


def _autosize(sheet: Worksheet, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        max_len = len(str(label))
        column_letter = get_column_letter(col_idx)
        for cell in sheet[column_letter]:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        sheet.column_dimensions[column_letter].width = min(max_len + 2, 40)
